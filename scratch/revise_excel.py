import openpyxl
import pandas as pd
import numpy as np

base_dir = "/Users/ruotingyang/Desktop/Projects/Meta subtype"
excel_path = f"{base_dir}/Meta subtype  Antigravity/workshop/source/Bridged metabolomics data_subtype analysis_2026.4.22_NEW CONTROLS-2.xlsx"
epi_path = f"{base_dir}/EpiagePTSD.csv"
ptsd_path = f"{base_dir}/allPTSDsamples 20170303.txt"

print("Loading external demographic references...")

# 1. Load Epiage
df_epi = pd.read_csv(epi_path)
df_epi['Extracted_ID'] = df_epi.iloc[:, 0].astype(str).str.extract(r'^(\d+)')[0]
epi_ages = df_epi.groupby('Extracted_ID')['Age'].mean().to_dict()

# 2. Manual Age corrections
manual_ages = {
    '202060': 30,
    '202210': 41,
    '211025': 31,
    '211028': 24,
    '212061': 29
}

# 3. Load allPTSDsamples for BMI
df_ptsd = pd.read_csv(ptsd_path, sep='\t')
df_ptsd['Extracted_ID'] = df_ptsd['Name'].astype(str).str.extract(r'^(\d+)')[0]
ptsd_bmi = df_ptsd.groupby('Extracted_ID')['BMI'].mean().to_dict()

print("Loading Excel workbook (this may take a minute due to file size)...")
wb = openpyxl.load_workbook(excel_path)
ws = wb['demographics']

# Find column indices (1-based)
header_row = 1
id_col, age_col, bmi_col, cohort_col = None, None, None, None

for cell in ws[header_row]:
    val = cell.value
    if val == 'ID': id_col = cell.column
    elif val == 'Age': age_col = cell.column
    elif val == 'BMI': bmi_col = cell.column
    elif val == 'Cohort': cohort_col = cell.column

if not all([id_col, age_col, bmi_col, cohort_col]):
    raise ValueError("Could not find all required columns in demographics sheet!")

age_updates = 0
bmi_updates = 0

print("Applying updates...")
for row in range(2, ws.max_row + 1):
    id_val = ws.cell(row=row, column=id_col).value
    if id_val is None:
        continue
    
    id_str = str(id_val).split('.')[0] # handle float parsing just in case
    
    # -- Age Update Logic --
    # 1. Manual first
    # 2. Then Epiage
    new_age = None
    if id_str in manual_ages:
        new_age = manual_ages[id_str]
    elif id_str in epi_ages:
        new_age = epi_ages[id_str]
        
    if new_age is not None:
        old_age = ws.cell(row=row, column=age_col).value
        # Use round(new_age, 2) or something if needed, but mean might be float
        if old_age != new_age and not (pd.isna(old_age) and pd.isna(new_age)):
            # Special case for NaN
            if old_age is None or old_age != new_age:
                ws.cell(row=row, column=age_col).value = float(new_age)
                age_updates += 1

    # -- BMI Update Logic --
    # Update BMI for SBC cohort OR if BMI is missing
    cohort_val = ws.cell(row=row, column=cohort_col).value
    old_bmi = ws.cell(row=row, column=bmi_col).value
    
    if id_str in ptsd_bmi:
        new_bmi = ptsd_bmi[id_str]
        if pd.notna(new_bmi):
            if cohort_val == 'SBC' or old_bmi is None:
                if old_bmi is None or abs(float(old_bmi) - float(new_bmi)) > 0.001:
                    ws.cell(row=row, column=bmi_col).value = float(new_bmi)
                    bmi_updates += 1

print(f"Updates completed: {age_updates} Age records, {bmi_updates} BMI records.")
print("Saving workbook...")
wb.save(excel_path)
print("Done!")
